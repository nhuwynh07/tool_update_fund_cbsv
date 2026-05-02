"""
statement_reader.py – Đọc file sao kê ngân hàng (Techcombank) thành DataFrame.
"""
import re

import pandas as pd
from openpyxl import load_workbook

from text_utils import norm_text, parse_money, parse_date
from config import DATA_START_ROW, NAME_COLS, TT_COL


# ── Tìm dòng tiêu đề trong sao kê ───────────────────────────────────────────

def find_statement_header_row(ws) -> int:
    """Quét từng dòng, tìm dòng có 'Ngày giao dịch' và 'Có TKTT/Credit'."""
    for row in range(1, ws.max_row + 1):
        joined = " ".join(
            norm_text(ws.cell(row, c).value)
            for c in range(1, ws.max_column + 1)
        )
        if (
            ("ngay giao dich" in joined or "transaction date" in joined)
            and ("co tktt" in joined or "credit" in joined)
        ):
            return row
    raise ValueError("Không tìm thấy dòng tiêu đề sao kê có Ngày giao dịch/Credit.")


def map_statement_columns(ws, header_row: int) -> dict:
    """Ánh xạ tên cột tiêu đề → chỉ số cột (int)."""
    cols: dict = {}
    for cell in ws[header_row]:
        h = norm_text(cell.value)
        if not h:
            continue
        if "ngay giao dich" in h or "transaction date" in h:
            cols["date"] = cell.column
        elif "nh doi tac" in h or "remitter bank" in h:
            cols["bank"] = cell.column
        elif "doi tac" in h or ("remitter" in h and "bank" not in h):
            cols["remitter"] = cell.column
        elif "dien giai" in h or "details" in h:
            cols["details"] = cell.column
        elif "so but toan" in h or "transaction no" in h:
            cols["txn_no"] = cell.column
        elif "co tktt" in h or h == "credit" or ("credit" in h and "co" in h):
            cols["credit"] = cell.column
        elif "no tktt" in h or h == "debit" or ("debit" in h and "no" in h):
            cols["debit"] = cell.column

    # Fallback: nếu norm_text không nhận ra, thử raw text
    if not cols:
        for cell in ws[header_row]:
            raw = str(cell.value or "").lower().replace("\n", " ")
            if "ng" in raw and "giao d" in raw:
                cols["date"] = cell.column
            elif "di" in raw and "n gi" in raw:
                cols["details"] = cell.column
            elif "c" in raw and "tktt" in raw and "debit" not in raw and "n" not in raw:
                cols["credit"] = cell.column

    missing = {"date", "details", "credit"} - set(cols)
    if missing:
        raise ValueError(f"Thiếu cột sao kê: {missing}. Cột tìm được: {cols}")
    return cols


# ── Lọc thông tin người nhận khỏi nội dung ──────────────────────────────────

def strip_recipient_info(value: str) -> str:
    """Xoá thông tin người nhận (tên thu quỹ, số TK) khỏi chuỗi tìm kiếm."""
    text = norm_text(value)
    text = re.sub(
        r"\bt\s*oi\s+\d[\d\s]*\s+[a-z\s]{0,100}?\s+tai\s+tech\s*combank\b",
        " ", text,
    )
    text = re.sub(
        r"\btoi\s+\d[\d\s]*\s+[a-z\s]{0,100}?\s+tai\s+tech\s*combank\b",
        " ", text,
    )
    text = text.replace("le thi nhu quynh", " ")
    text = text.replace("le thi nhu quy nh", " ")
    text = text.replace("3832862685", " ")
    return norm_text(text)


# ── Đọc toàn bộ sao kê ──────────────────────────────────────────────────────

def read_statement(path) -> pd.DataFrame:
    """
    Đọc file sao kê Excel, trả về DataFrame các giao dịch phát sinh Có (credit > 0).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = find_statement_header_row(ws)
    cols = map_statement_columns(ws, header_row)
    rows = []
    blank_count = 0

    for r in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(r, cols["date"]).value
        date     = parse_date(raw_date)
        details  = str(ws.cell(r, cols["details"]).value or "")
        credit   = parse_money(ws.cell(r, cols["credit"]).value)

        row_text = norm_text(
            " ".join(str(ws.cell(r, c).value or "") for c in cols.values())
        )
        if (
            "cong doanh so" in row_text
            or "total volume" in row_text
            or "so du cuoi ky" in row_text
        ):
            break

        if not raw_date and not details and credit == 0:
            blank_count += 1
            if blank_count >= 8:
                break
            continue
        blank_count = 0

        if pd.isna(date) or credit <= 0:
            continue

        remitter    = str(ws.cell(r, cols.get("remitter", 1)).value or "")
        bank        = str(ws.cell(r, cols.get("bank", 1)).value or "")
        txn_no      = str(ws.cell(r, cols.get("txn_no", 1)).value or "")
        search_text = strip_recipient_info(f"{remitter} {details}")

        rows.append({
            "excel_row":   r,
            "date":        date,
            "amount":      credit,
            "remitter":    remitter,
            "bank":        bank,
            "details":     details,
            "txn_no":      txn_no,
            "search_text": search_text,
        })

    return pd.DataFrame(rows)


# ── Đọc danh sách thành viên từ sheet ───────────────────────────────────────

def read_members(ws) -> list:
    """
    Đọc danh sách thành viên (TT, tên) từ sheet tổng hợp.
    Trả về list dict: {row, tt, name, name_norm}.
    """
    import re as _re
    from text_utils import norm_text as _norm

    members = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        tt = ws.cell(r, TT_COL).value
        if not isinstance(tt, int):
            if isinstance(tt, float) and tt.is_integer():
                tt = int(tt)
            elif isinstance(tt, str) and tt.strip().isdigit():
                tt = int(tt.strip())
            else:
                continue
        name = " ".join(str(ws.cell(r, c).value or "").strip() for c in NAME_COLS).strip()
        name = _re.sub(r"\s+", " ", name)
        if name:
            members.append({"row": r, "tt": tt, "name": name, "name_norm": _norm(name)})
    return members
