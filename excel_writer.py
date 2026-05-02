"""
excel_writer.py – Cập nhật dữ liệu vào sheet Excel và ghi sheet log đối soát.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from text_utils import norm_text, parse_money
from config import (
    LOG_SHEET,
    QUY_MONTH_COLS, QUY_STANDARD_MONTH_AMOUNT,
    PHAT_TOTAL_COL,
)


# ── Cập nhật sheet Quỹ ───────────────────────────────────────────────────────

def refresh_quy_formulas(ws, row: int) -> None:
    """Ghi lại công thức tổng quý và tổng năm cho dòng thành viên."""
    ws.cell(row, 8).value  = f"=SUM(E{row}:G{row})"
    ws.cell(row, 12).value = f"=SUM(I{row}:K{row})"
    ws.cell(row, 16).value = f"=SUM(M{row}:O{row})"
    ws.cell(row, 20).value = f"=SUM(Q{row}:S{row})"
    ws.cell(row, 21).value = f"=SUM(H{row}+L{row}+P{row}+T{row})"


def set_quy_month(ws, row: int, month: int) -> str:
    """
    Điền giá trị tháng quỹ chuẩn vào ô tương ứng.
    Trả về mô tả hành động: 'giữ nguyên' | 'cập nhật' | 'ghi đè <cũ>'.
    """
    cell = ws.cell(row, QUY_MONTH_COLS[month])
    old  = parse_money(cell.value)
    if old == QUY_STANDARD_MONTH_AMOUNT:
        return "giữ nguyên"
    cell.value = QUY_STANDARD_MONTH_AMOUNT
    return "cập nhật" if old == 0 else f"ghi đè {old}"


# ── Cập nhật sheet Phạt ──────────────────────────────────────────────────────

def refresh_phat_formula(ws, row: int) -> None:
    """Ghi lại công thức tổng phạt cho dòng thành viên."""
    ws.cell(row, PHAT_TOTAL_COL).value = f"=SUM(E{row}:J{row})"


def update_phat_cell(ws, row: int, col: int, amount: int) -> str:
    """
    Cộng dồn số tiền phạt vào ô tương ứng.
    Trả về mô tả hành động: 'giữ nguyên' | 'cập nhật' | 'cộng thêm, cũ=<cũ>'.
    """
    cell = ws.cell(row, col)
    old  = parse_money(cell.value)
    if old == amount:
        return "giữ nguyên"
    if old == 0:
        cell.value = amount
        return "cập nhật"
    cell.value = old + amount
    return f"cộng thêm, cũ={old}"


# ── Đọc/ghi sheet Log ────────────────────────────────────────────────────────

def read_logged_txn_ids(wb) -> set:
    """
    Đọc tập hợp 'Số bút toán' đã có trong sheet log cũ,
    dùng để bỏ qua các giao dịch đã xử lý lần trước.
    """
    if LOG_SHEET not in wb.sheetnames:
        return set()
    ws = wb[LOG_SHEET]
    headers = [norm_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    try:
        col_idx = headers.index(norm_text("Số bút toán")) + 1
    except ValueError:
        return set()
    seen = set()
    for r in range(2, ws.max_row + 1):
        value = str(ws.cell(r, col_idx).value or "").strip()
        if value:
            seen.add(value)
    return seen


def write_log_sheet(wb, logs: list) -> None:
    """
    Tạo (hoặc ghi đè) sheet log với toàn bộ kết quả đối soát.
    Áp dụng định dạng bảng: border, wrap text, tiêu đề in đậm nền xanh.
    """
    if LOG_SHEET in wb.sheetnames:
        del wb[LOG_SHEET]
    ws = wb.create_sheet(LOG_SHEET)

    headers = [
        "Ngày", "Dòng sao kê", "Số bút toán", "Số tiền",
        "Nội dung", "Loại", "Tên khớp", "Điểm",
        "Sheet", "Ô/Cột cập nhật", "Trạng thái",
    ]
    ws.append(headers)
    for item in logs:
        ws.append([item.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)

    widths = {
        "A": 12, "B": 12, "C": 22, "D": 14, "E": 70,
        "F": 14, "G": 28, "H": 10, "I": 18, "J": 24, "K": 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
