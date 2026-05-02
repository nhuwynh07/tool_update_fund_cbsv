"""
main.py – Điểm khởi chạy chương trình đối soát thu chi bộ.

Luồng xử lý:
    1. Đọc sao kê ngân hàng  (statement_reader)
    2. Đọc danh sách thành viên từ file tổng hợp
    3. Với mỗi giao dịch:
       a. Phân loại: quỹ / phạt / bỏ qua / không xác định  (matcher)
       b. Khớp tên thành viên                               (matcher)
       c. Cập nhật ô Excel tương ứng                        (excel_writer)
       d. Ghi log
    4. Lưu file kết quả + in tóm tắt

Hàm run_core() hỗ trợ cả đường dẫn file và BytesIO,
dùng được từ CLI lẫn giao diện Streamlit.
"""
import io

import pandas as pd
from openpyxl import load_workbook

from config import (
    SAO_KE_PATH, TONG_HOP_PATH, OUTPUT_PATH,
    SHEET_QUY, SHEET_PHAT,
    QUY_MONTH_COLS,
)
from statement_reader import read_statement, read_members
from matcher import classify_transaction, match_member, get_quy_months
from excel_writer import (
    read_logged_txn_ids, write_log_sheet,
    set_quy_month, refresh_quy_formulas,
    update_phat_cell, refresh_phat_formula,
)


# ── Hàm xử lý core (dùng chung cho CLI và Streamlit) ────────────────────────

def run_core(sao_ke_source, tong_hop_source) -> dict:
    """
    Thực hiện toàn bộ quy trình đối soát.

    Parameters
    ----------
    sao_ke_source   : str | Path | BytesIO  – file sao kê ngân hàng
    tong_hop_source : str | Path | BytesIO  – file báo cáo tổng hợp

    Returns
    -------
    dict với các key:
        logs          – pd.DataFrame toàn bộ log giao dịch
        excel_bytes   – bytes của file Excel đã cập nhật
        workbook      – openpyxl Workbook (đã lưu vào excel_bytes)
        members_quy   – list thành viên từ sheet Quỹ
        n_transactions – số giao dịch Có đọc được từ sao kê
    """
    df      = read_statement(sao_ke_source)
    wb      = load_workbook(tong_hop_source)
    ws_quy  = wb[SHEET_QUY]
    ws_phat = wb[SHEET_PHAT]

    members_quy  = read_members(ws_quy)
    members_phat = read_members(ws_phat)
    seen_txn_ids = read_logged_txn_ids(wb)
    logs: list   = []

    for _, row in df.iterrows():
        tx = row.to_dict()
        tx_type, target_col, type_note = classify_transaction(tx)

        log = {
            "Ngày":           tx["date"].strftime("%d/%m/%Y"),
            "Dòng sao kê":    tx["excel_row"],
            "Số bút toán":    tx["txn_no"],
            "Số tiền":        tx["amount"],
            "Nội dung":       tx["details"],
            "Loại":           tx_type,
            "Tên khớp":       "",
            "Điểm":           "",
            "Sheet":          "",
            "Ô/Cột cập nhật": "",
            "Trạng thái":     "",
        }

        if tx["txn_no"] and tx["txn_no"] in seen_txn_ids:
            log["Trạng thái"] = "Bỏ qua vì số bút toán đã có trong log cũ"
            logs.append(log)
            continue

        if tx_type in ("ignore", "unknown"):
            log["Trạng thái"] = type_note
            logs.append(log)
            continue

        members   = members_quy  if tx_type == "quy" else members_phat
        target_ws = ws_quy       if tx_type == "quy" else ws_phat
        member, score, match_note = match_member(tx, members)
        log["Điểm"] = round(score, 1)

        if member is None:
            log["Trạng thái"] = match_note
            logs.append(log)
            continue

        log["Tên khớp"] = member["name"]
        log["Sheet"]    = target_ws.title

        if tx_type == "quy":
            months  = get_quy_months(tx)
            actions = []
            for month in months:
                action = set_quy_month(ws_quy, member["row"], month)
                actions.append(f"T{month}:{action}")
            refresh_quy_formulas(ws_quy, member["row"])
            log["Ô/Cột cập nhật"] = ", ".join(
                ws_quy.cell(member["row"], QUY_MONTH_COLS[m]).coordinate
                for m in months
            )
            log["Trạng thái"] = f"{type_note}; " + ", ".join(actions)
        else:
            action = update_phat_cell(ws_phat, member["row"], target_col, tx["amount"])
            refresh_phat_formula(ws_phat, member["row"])
            log["Ô/Cột cập nhật"] = ws_phat.cell(member["row"], target_col).coordinate
            log["Trạng thái"]     = f"{type_note}; {action}"

        logs.append(log)

    write_log_sheet(wb, logs)

    # Lưu workbook vào bộ nhớ (BytesIO) thay vì ghi thẳng ra đĩa
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return {
        "logs":          pd.DataFrame(logs),
        "excel_bytes":   buf.getvalue(),
        "workbook":      wb,
        "members_quy":   members_quy,
        "n_transactions": len(df),
    }


# ── CLI entry point ───────────────────────────────────────────────────────────

def run() -> None:
    """Chạy từ command line, đọc file từ config, lưu kết quả ra đĩa."""
    import sys
    import io as _io

    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    result = run_core(SAO_KE_PATH, TONG_HOP_PATH)

    with open(OUTPUT_PATH, "wb") as f:
        f.write(result["excel_bytes"])

    summary = (
        result["logs"]
        .groupby(["Loại", "Trạng thái"], dropna=False)
        .size()
        .reset_index(name="Số dòng")
    )
    print(f"Đã lưu: {OUTPUT_PATH}")
    print(f"Đọc {result['n_transactions']} giao dịch có phát sinh Có từ sao kê.")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    run()
